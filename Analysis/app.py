from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_file
)

import numpy as np
import io
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import (
    get_stock_data,
    get_available_tickers,
    get_database_status
)

from calculations import (
    calculate_all_periods,
    calculate_groups_analysis
)


app = Flask(__name__)


# ============================================================
# HELPERS
# ============================================================

def clean_value(value):
    if value is None:
        return None

    if isinstance(value, (np.integer, int)):
        return int(value)

    if isinstance(value, (np.floating, float)):
        if np.isnan(value) or not np.isfinite(value):
            return None
        return float(value)

    return value


def clean_dict(data):
    if isinstance(data, dict):
        return {
            str(key): clean_dict(value)
            for key, value in data.items()
        }

    if isinstance(data, list):
        return [
            clean_dict(value)
            for value in data
        ]

    return clean_value(data)


# ============================================================
# MAIN PAGE
# ============================================================

@app.route("/")
def index():
    try:
        tickers = get_available_tickers()
    except Exception as e:
        print("Database error:", e)
        tickers = []

    return render_template(
        "index.html",
        tickers=tickers
    )


# ============================================================
# SINGLE ASSET CALCULATION
# ============================================================

@app.route("/api/calculate", methods=["POST"])
def calculate():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Пустой запрос."}), 400

        stock = str(data.get("stock", "")).strip().upper()
        benchmark = str(data.get("benchmark", "")).strip().upper()
        benchmark_move = data.get("benchmark_move", 0.01)
        asset_move = data.get("asset_move")

        try:
            benchmark_move = float(benchmark_move)
        except (TypeError, ValueError):
            return jsonify({"error": "Некорректное движение benchmark."}), 400

        if asset_move is not None:
            try:
                asset_move = float(asset_move)
            except (TypeError, ValueError):
                return jsonify({"error": "Некорректное движение акции."}), 400

        if not stock:
            return jsonify({"error": "Не указана акция."}), 400

        if not benchmark:
            return jsonify({"error": "Не указан benchmark."}), 400

        if stock == benchmark:
            return jsonify({"error": "Акции должны быть разными."}), 400

        df = get_stock_data(stock, benchmark)
        if df is None or df.empty:
            return jsonify({"error": "Данные не найдены."}), 404

        results, _ = calculate_all_periods(
            df,
            stock,
            benchmark,
            benchmark_move=benchmark_move,
            asset_move=asset_move
        )

        return jsonify({
            "stock": stock,
            "benchmark": benchmark,
            "benchmark_move": benchmark_move,
            "asset_move": asset_move,
            "results": clean_dict(results),
            "chart": []
        })

    except Exception as e:
        print("Calculation error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# GROUP / BATCH ANALYSIS (UP TO 300+ TICKERS)
# ============================================================

# ============================================================
# GROUP / BATCH ANALYSIS (UP TO 300+ TICKERS & MULTI-BENCHMARK)
# ============================================================

@app.route("/api/calculate-groups", methods=["POST"])
def calculate_groups():
    import re
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Пустой запрос."}), 400

        groups = data.get("groups")
        benchmark_raw = data.get("benchmarks") or data.get("benchmark") or data.get("reference")
        period = data.get("period") or "1y"
        benchmark_move = data.get("benchmark_move", 0.01)
        benchmark_moves = data.get("benchmark_moves", {})
        asset_moves = data.get("asset_moves", {})

        try:
            benchmark_move = float(benchmark_move)
        except (TypeError, ValueError):
            return jsonify({"error": "Некорректное движение benchmark."}), 400

        if not groups or not isinstance(groups, dict):
            return jsonify({"error": "Добавьте хотя бы одну группу с тикерами."}), 400

        if not benchmark_raw:
            return jsonify({"error": "Не указан Market Asset (Benchmark)."}), 400

        # Parse multiple benchmarks
        benchmarks_list = []
        if isinstance(benchmark_raw, (list, tuple, set)):
            benchmarks_list = [str(b).strip().upper() for b in benchmark_raw if str(b).strip()]
        elif isinstance(benchmark_raw, str):
            benchmarks_list = [b.strip().upper() for b in re.split(r"[\s,;]+", benchmark_raw) if b.strip()]

        clean_benchmarks = []
        for bm in benchmarks_list:
            if bm and bm not in clean_benchmarks:
                clean_benchmarks.append(bm)

        if not clean_benchmarks:
            return jsonify({"error": "Не указан валидный Market Asset (Benchmark)."}), 400

        # Clean Groups & Tickers
        cleaned_groups = {}
        all_requested_tickers = []

        for group_name, tickers in groups.items():
            if not isinstance(tickers, list):
                continue

            clean_tickers = []
            for ticker in tickers:
                if not isinstance(ticker, str):
                    continue
                ticker = ticker.strip().upper()
                if ticker and ticker not in clean_benchmarks and ticker not in clean_tickers:
                    clean_tickers.append(ticker)
                    if ticker not in all_requested_tickers:
                        all_requested_tickers.append(ticker)

            if clean_tickers:
                name = str(group_name).strip() or "Group"
                cleaned_groups[name] = clean_tickers

        if not cleaned_groups or not all_requested_tickers:
            return jsonify({"error": "Нет валидных тикеров для анализа."}), 400

        all_symbols = list(dict.fromkeys(all_requested_tickers + clean_benchmarks))

        # Fetch Data from Database
        df = get_stock_data(*all_symbols)
        if df is None or df.empty:
            return jsonify({"error": "Данные для указанных тикеров не найдены в базе данных."}), 404

        available_tickers = set(
            df["ticker"].astype(str).str.strip().str.upper().unique()
        )

        missing_benchmarks = [bm for bm in clean_benchmarks if bm not in available_tickers]
        if missing_benchmarks:
            return jsonify({
                "error": f"Market Asset '{', '.join(missing_benchmarks)}' не найден в базе данных. Проверьте правильность тикеров."
            }), 404

        # Check missing tickers (graceful handling for 300+ batch)
        missing_tickers = [
            ticker for ticker in all_requested_tickers
            if ticker not in available_tickers
        ]

        # Filter groups to only available tickers
        active_groups = {}
        for g_name, t_list in cleaned_groups.items():
            valid_t = [t for t in t_list if t in available_tickers]
            if valid_t:
                active_groups[g_name] = valid_t

        if not active_groups:
            return jsonify({
                "error": "Ни один из указанных тикеров не найден в базе данных.",
                "missing": missing_tickers
            }), 404

        # Clean Asset Moves
        if not isinstance(asset_moves, dict):
            asset_moves = {}

        cleaned_asset_moves = {}
        for ticker, move in asset_moves.items():
            ticker = str(ticker).strip().upper()
            try:
                move = float(move)
            except (TypeError, ValueError):
                continue
            if np.isfinite(move):
                cleaned_asset_moves[ticker] = move

        # Run Vectorized Multi-Benchmark Group Analysis
        analysis = calculate_groups_analysis(
            df=df,
            groups=active_groups,
            benchmarks=clean_benchmarks,
            period=period,
            benchmark_move=benchmark_move,
            benchmark_moves=benchmark_moves if isinstance(benchmark_moves, dict) else None,
            asset_moves=cleaned_asset_moves
        )

        ignored_tickers = analysis.get("ignored_tickers", {})
        partial_tickers = analysis.get("partial_tickers", {})
        warning_parts = []
        if missing_tickers:
            warning_parts.append(
                f"Пропущено {len(missing_tickers)} тикеров (нет в БД): {', '.join(missing_tickers[:8])}"
                + (f" и ещё {len(missing_tickers) - 8}..." if len(missing_tickers) > 8 else "")
            )
        if ignored_tickers:
            ignored_list = list(ignored_tickers.keys()) if isinstance(ignored_tickers, dict) else list(ignored_tickers)
            warning_parts.append(
                f"Не удалось рассчитать {len(ignored_list)} тикеров (< 2 наблюдений): {', '.join(ignored_list[:8])}"
                + (f" и ещё {len(ignored_list) - 8}..." if len(ignored_list) > 8 else "")
            )
        if partial_tickers:
            partial_list = list(partial_tickers.keys())
            warning_parts.append(
                f"⚠️ {len(partial_list)} тикеров имеют историю меньше периода {period} и подсвечены в таблице: {', '.join(partial_list[:6])}"
                + (f" и ещё {len(partial_list) - 6}..." if len(partial_list) > 6 else "")
            )
        warning_message = " | ".join(warning_parts) if warning_parts else None

        response = {
            "groups": active_groups,
            "benchmarks": clean_benchmarks,
            "primary_benchmark": clean_benchmarks[0],
            "benchmark": clean_benchmarks[0],
            "reference": clean_benchmarks[0],
            "benchmark_moves": clean_dict(analysis.get("benchmark_moves", {})),
            "benchmark_move": benchmark_move,
            "asset_moves": cleaned_asset_moves,
            "tickers": analysis.get("tickers", []),
            "all_symbols": analysis.get("all_symbols", []),
            "period": analysis.get("period", period),
            "period_start": analysis.get("period_start"),
            "period_end": analysis.get("period_end"),
            "correlation": clean_dict(analysis.get("correlation", {})),
            "beta": clean_dict(analysis.get("beta", {})),
            "benchmark_beta": clean_dict(analysis.get("benchmark_beta", {})),
            "expected_range": clean_dict(analysis.get("expected_range", {})),
            "benchmarks_data": clean_dict(analysis.get("benchmarks_data", {})),
            "group_summary": clean_dict(analysis.get("group_summary", {})),
            "missing_tickers": missing_tickers,
            "ignored_tickers": clean_dict(ignored_tickers),
            "partial_tickers": clean_dict(partial_tickers),
            "warning": warning_message,
            "total_requested": len(all_requested_tickers),
            "total_analyzed": len(analysis.get("tickers", []))
        }

        return jsonify(response)

    except Exception as e:
        print("Group calculation error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# EXPORT FILTERED CORRELATION RESULTS TO EXCEL (.XLSX)
# ============================================================

@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    try:
        data = request.get_json() or {}

        rows = data.get("rows", [])
        benchmarks = data.get("benchmarks", [])
        benchmark = str(data.get("benchmark", "")).strip().upper() or "BENCHMARK"
        benchmarks_data = data.get("benchmarks_data", {})
        period = str(data.get("period", "1y"))
        filter_label = str(data.get("filter_label", "All Assets")).strip()
        is_multi_benchmark = bool(data.get("is_multi_benchmark", False) or (isinstance(benchmarks, list) and len(benchmarks) > 1))

        if not isinstance(rows, list) or not rows:
            return jsonify({"error": "Нет результатов для экспорта в Excel."}), 400

        if not benchmarks:
            benchmarks = [benchmark]

        wb = openpyxl.Workbook()

        # Styles definition
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_fill_accent = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_thin = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        # Correlation fills
        fill_high_pos = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # soft green
        fill_mod_pos = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # light green
        fill_high_neg = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # soft red
        fill_mod_neg = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")  # light red

        # Format helper for correlation cells
        def format_corr_cell(cell, corr):
            if corr is not None:
                cell.number_format = "0.000"
                if corr >= 0.50:
                    cell.fill = fill_high_pos
                    cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
                elif corr >= 0.20:
                    cell.fill = fill_mod_pos
                    cell.font = Font(name="Calibri", size=10, color="15803D")
                elif corr <= -0.50:
                    cell.fill = fill_high_neg
                    cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")
                elif corr < 0:
                    cell.fill = fill_mod_neg
                    cell.font = Font(name="Calibri", size=10, color="B91C1C")

        # ========================================================
        # SHEET 1: Main View (Multi-Benchmark Comparison OR Single Benchmark)
        # ========================================================
        ws1 = wb.active

        if is_multi_benchmark and len(benchmarks) > 1:
            ws1.title = "Benchmark Comparison"

            headers = ["Ticker"]
            for bm in benchmarks:
                headers.extend([f"Corr ({bm})", f"Beta ({bm})", f"Exp Move ({bm})"])
            headers.extend(["Actual Move", "Daily Volatility", "Observations", "Period"])

            ws1.append(headers)
            ws1.row_dimensions[1].height = 26

            for col_num in range(1, len(headers) + 1):
                cell = ws1.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = cell_border

            # Populate Rows
            for row_idx, r in enumerate(rows, start=2):
                ticker = str(r.get("ticker", "")).strip().upper()
                row_data = [ticker]

                for bm in benchmarks:
                    bm_corr = clean_value(r.get(f"correlation_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("correlation"))
                    bm_beta = clean_value(r.get(f"beta_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("beta"))
                    bm_exp = clean_value(r.get(f"expected_move_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("expected_move"))
                    row_data.extend([bm_corr, bm_beta, bm_exp])

                act_move = clean_value(r.get("actual_move"))
                vol = clean_value(r.get("volatility"))
                obs = clean_value(r.get("observations"))

                row_data.extend([act_move, vol, obs, period])
                ws1.append(row_data)
                ws1.row_dimensions[row_idx].height = 20

                # Format Cells
                col_c = 1
                c_ticker = ws1.cell(row=row_idx, column=col_c)
                c_ticker.alignment = Alignment(horizontal="center", vertical="center")
                c_ticker.font = Font(name="Calibri", size=10, bold=True)
                c_ticker.border = cell_border
                col_c += 1

                for bm in benchmarks:
                    # Corr
                    c_corr = ws1.cell(row=row_idx, column=col_c)
                    c_corr.alignment = Alignment(horizontal="right", vertical="center")
                    c_corr.border = cell_border
                    bm_corr = clean_value(r.get(f"correlation_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("correlation"))
                    format_corr_cell(c_corr, bm_corr)
                    col_c += 1

                    # Beta
                    c_beta = ws1.cell(row=row_idx, column=col_c)
                    c_beta.alignment = Alignment(horizontal="right", vertical="center")
                    c_beta.border = cell_border
                    bm_beta = clean_value(r.get(f"beta_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("beta"))
                    if bm_beta is not None:
                        c_beta.number_format = "0.000"
                    col_c += 1

                    # Exp Move
                    c_exp = ws1.cell(row=row_idx, column=col_c)
                    c_exp.alignment = Alignment(horizontal="right", vertical="center")
                    c_exp.border = cell_border
                    bm_exp = clean_value(r.get(f"expected_move_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("expected_move"))
                    if bm_exp is not None:
                        c_exp.number_format = "0.00%"
                    col_c += 1

                # Actual move
                c_act = ws1.cell(row=row_idx, column=col_c)
                c_act.alignment = Alignment(horizontal="right", vertical="center")
                c_act.border = cell_border
                if act_move is not None:
                    c_act.number_format = "0.00%"
                col_c += 1

                # Volatility
                c_vol = ws1.cell(row=row_idx, column=col_c)
                c_vol.alignment = Alignment(horizontal="right", vertical="center")
                c_vol.border = cell_border
                if vol is not None:
                    c_vol.number_format = "0.00%"
                col_c += 1

                # Obs
                c_obs = ws1.cell(row=row_idx, column=col_c)
                c_obs.alignment = Alignment(horizontal="center", vertical="center")
                c_obs.border = cell_border
                if obs is not None:
                    c_obs.number_format = "#,##0"
                col_c += 1

                # Period
                c_per = ws1.cell(row=row_idx, column=col_c)
                c_per.alignment = Alignment(horizontal="center", vertical="center")
                c_per.border = cell_border

            ws1.freeze_panes = "A2"
            ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

            # Auto column widths
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                header_text = str(headers[col_idx - 1])
                ws1.column_dimensions[col_letter].width = max(len(header_text) + 4, 13)

            # ========================================================
            # ADDITIONAL SHEETS FOR EACH INDIVIDUAL BENCHMARK
            # ========================================================
            for bm in benchmarks:
                sheet_title = f"{bm} Analysis"[:31]
                ws_bm = wb.create_sheet(title=sheet_title)

                bm_headers = [
                    "Ticker",
                    "Correlation",
                    "Beta",
                    "Expected Move",
                    "Actual Move",
                    "Difference",
                    "Normalized Diff",
                    "Daily Volatility",
                    "Residual Volatility",
                    "Observations",
                    "Benchmark",
                    "Period"
                ]
                ws_bm.append(bm_headers)
                ws_bm.row_dimensions[1].height = 26

                for col_num in range(1, len(bm_headers) + 1):
                    cell = ws_bm.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = cell_border

                for row_idx, r in enumerate(rows, start=2):
                    ticker = str(r.get("ticker", "")).strip().upper()
                    m = r.get("benchmarks_metrics", {}).get(bm, {}) if isinstance(r.get("benchmarks_metrics"), dict) else {}
                    corr = clean_value(r.get(f"correlation_{bm}", m.get("correlation", r.get("correlation"))))
                    beta_val = clean_value(r.get(f"beta_{bm}", m.get("beta", r.get("beta"))))
                    exp_move = clean_value(r.get(f"expected_move_{bm}", m.get("expected_move", r.get("expected_move"))))
                    act_move = clean_value(r.get("actual_move", m.get("actual_move")))
                    diff = clean_value(m.get("difference", r.get(f"difference_{bm}", r.get("difference"))))
                    norm = clean_value(m.get("normalized", r.get(f"normalized_{bm}", r.get("normalized"))))
                    vol = clean_value(r.get("volatility", m.get("volatility")))
                    res_vol = clean_value(m.get("residual_volatility", r.get(f"residual_volatility_{bm}", r.get("residual_volatility"))))
                    obs = clean_value(r.get("observations", m.get("observations")))

                    row_data = [
                        ticker,
                        corr,
                        beta_val,
                        exp_move,
                        act_move,
                        diff,
                        norm,
                        vol,
                        res_vol,
                        obs,
                        bm,
                        period
                    ]
                    ws_bm.append(row_data)
                    ws_bm.row_dimensions[row_idx].height = 20

                    # Styling
                    c_tk = ws_bm.cell(row=row_idx, column=1)
                    c_tk.alignment = Alignment(horizontal="center", vertical="center")
                    c_tk.font = Font(name="Calibri", size=10, bold=True)
                    c_tk.border = cell_border

                    c_cr = ws_bm.cell(row=row_idx, column=2)
                    c_cr.alignment = Alignment(horizontal="right", vertical="center")
                    c_cr.border = cell_border
                    format_corr_cell(c_cr, corr)

                    c_bt = ws_bm.cell(row=row_idx, column=3)
                    c_bt.alignment = Alignment(horizontal="right", vertical="center")
                    c_bt.border = cell_border
                    if beta_val is not None:
                        c_bt.number_format = "0.000"

                    c_ex = ws_bm.cell(row=row_idx, column=4)
                    c_ex.alignment = Alignment(horizontal="right", vertical="center")
                    c_ex.border = cell_border
                    if exp_move is not None:
                        c_ex.number_format = "0.00%"

                    c_ac = ws_bm.cell(row=row_idx, column=5)
                    c_ac.alignment = Alignment(horizontal="right", vertical="center")
                    c_ac.border = cell_border
                    if act_move is not None:
                        c_ac.number_format = "0.00%"

                    c_df = ws_bm.cell(row=row_idx, column=6)
                    c_df.alignment = Alignment(horizontal="right", vertical="center")
                    c_df.border = cell_border
                    if diff is not None:
                        c_df.number_format = "0.00%"
                        if diff > 0:
                            c_df.font = Font(name="Calibri", size=10, color="16A34A")
                        elif diff < 0:
                            c_df.font = Font(name="Calibri", size=10, color="DC2626")

                    c_nm = ws_bm.cell(row=row_idx, column=7)
                    c_nm.alignment = Alignment(horizontal="right", vertical="center")
                    c_nm.border = cell_border
                    if norm is not None:
                        c_nm.number_format = "0.00"

                    c_vl = ws_bm.cell(row=row_idx, column=8)
                    c_vl.alignment = Alignment(horizontal="right", vertical="center")
                    c_vl.border = cell_border
                    if vol is not None:
                        c_vl.number_format = "0.00%"

                    c_rv = ws_bm.cell(row=row_idx, column=9)
                    c_rv.alignment = Alignment(horizontal="right", vertical="center")
                    c_rv.border = cell_border
                    if res_vol is not None:
                        c_rv.number_format = "0.00%"

                    c_ob = ws_bm.cell(row=row_idx, column=10)
                    c_ob.alignment = Alignment(horizontal="center", vertical="center")
                    c_ob.border = cell_border
                    if obs is not None:
                        c_ob.number_format = "#,##0"

                    c_bm_c = ws_bm.cell(row=row_idx, column=11)
                    c_bm_c.alignment = Alignment(horizontal="center", vertical="center")
                    c_bm_c.border = cell_border

                    c_pr = ws_bm.cell(row=row_idx, column=12)
                    c_pr.alignment = Alignment(horizontal="center", vertical="center")
                    c_pr.border = cell_border

                ws_bm.freeze_panes = "A2"
                ws_bm.auto_filter.ref = f"A1:L{len(rows) + 1}"

                col_widths = {
                    "A": 12, "B": 15, "C": 12, "D": 16, "E": 15,
                    "F": 15, "G": 16, "H": 16, "I": 18, "J": 14,
                    "K": 13, "L": 11
                }
                for col_letter, width in col_widths.items():
                    ws_bm.column_dimensions[col_letter].width = width

        else:
            # Single benchmark view
            ws1.title = "Correlation Analysis"

            headers = [
                "Ticker",
                "Correlation",
                "Beta",
                "Expected Move",
                "Actual Move",
                "Difference",
                "Normalized Diff",
                "Daily Volatility",
                "Residual Volatility",
                "Observations",
                "Benchmark",
                "Period"
            ]
            ws1.append(headers)
            ws1.row_dimensions[1].height = 26

            for col_num in range(1, len(headers) + 1):
                cell = ws1.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = cell_border

            for row_idx, r in enumerate(rows, start=2):
                ticker = str(r.get("ticker", "")).strip().upper()
                corr = clean_value(r.get("correlation"))
                beta_val = clean_value(r.get("beta"))
                exp_move = clean_value(r.get("expected_move"))
                act_move = clean_value(r.get("actual_move"))
                diff = clean_value(r.get("difference"))
                norm = clean_value(r.get("normalized"))
                vol = clean_value(r.get("volatility"))
                res_vol = clean_value(r.get("residual_volatility"))
                obs = clean_value(r.get("observations"))

                row_data = [
                    ticker,
                    corr,
                    beta_val,
                    exp_move,
                    act_move,
                    diff,
                    norm,
                    vol,
                    res_vol,
                    obs,
                    benchmark,
                    period
                ]
                ws1.append(row_data)
                ws1.row_dimensions[row_idx].height = 20

                c_ticker = ws1.cell(row=row_idx, column=1)
                c_ticker.alignment = Alignment(horizontal="center", vertical="center")
                c_ticker.font = Font(name="Calibri", size=10, bold=True)
                c_ticker.border = cell_border

                c_corr = ws1.cell(row=row_idx, column=2)
                c_corr.alignment = Alignment(horizontal="right", vertical="center")
                c_corr.border = cell_border
                format_corr_cell(c_corr, corr)

                c_beta = ws1.cell(row=row_idx, column=3)
                c_beta.alignment = Alignment(horizontal="right", vertical="center")
                c_beta.border = cell_border
                if beta_val is not None:
                    c_beta.number_format = "0.000"

                c_exp = ws1.cell(row=row_idx, column=4)
                c_exp.alignment = Alignment(horizontal="right", vertical="center")
                c_exp.border = cell_border
                if exp_move is not None:
                    c_exp.number_format = "0.00%"

                c_act = ws1.cell(row=row_idx, column=5)
                c_act.alignment = Alignment(horizontal="right", vertical="center")
                c_act.border = cell_border
                if act_move is not None:
                    c_act.number_format = "0.00%"

                c_diff = ws1.cell(row=row_idx, column=6)
                c_diff.alignment = Alignment(horizontal="right", vertical="center")
                c_diff.border = cell_border
                if diff is not None:
                    c_diff.number_format = "0.00%"
                    if diff > 0:
                        c_diff.font = Font(name="Calibri", size=10, color="16A34A")
                    elif diff < 0:
                        c_diff.font = Font(name="Calibri", size=10, color="DC2626")

                c_norm = ws1.cell(row=row_idx, column=7)
                c_norm.alignment = Alignment(horizontal="right", vertical="center")
                c_norm.border = cell_border
                if norm is not None:
                    c_norm.number_format = "0.00"

                c_vol = ws1.cell(row=row_idx, column=8)
                c_vol.alignment = Alignment(horizontal="right", vertical="center")
                c_vol.border = cell_border
                if vol is not None:
                    c_vol.number_format = "0.00%"

                c_res = ws1.cell(row=row_idx, column=9)
                c_res.alignment = Alignment(horizontal="right", vertical="center")
                c_res.border = cell_border
                if res_vol is not None:
                    c_res.number_format = "0.00%"

                c_obs = ws1.cell(row=row_idx, column=10)
                c_obs.alignment = Alignment(horizontal="center", vertical="center")
                c_obs.border = cell_border
                if obs is not None:
                    c_obs.number_format = "#,##0"

                c_bm = ws1.cell(row=row_idx, column=11)
                c_bm.alignment = Alignment(horizontal="center", vertical="center")
                c_bm.border = cell_border

                c_per = ws1.cell(row=row_idx, column=12)
                c_per.alignment = Alignment(horizontal="center", vertical="center")
                c_per.border = cell_border

            ws1.freeze_panes = "A2"
            ws1.auto_filter.ref = f"A1:L{len(rows) + 1}"

            col_widths = {
                "A": 12, "B": 15, "C": 12, "D": 16, "E": 15,
                "F": 15, "G": 16, "H": 16, "I": 18, "J": 14,
                "K": 13, "L": 11
            }
            for col_letter, width in col_widths.items():
                ws1.column_dimensions[col_letter].width = width

        # ========================================================
        # SHEET: Summary & Export Metadata
        # ========================================================
        ws_summary = wb.create_sheet(title="Summary & Metadata")
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.append(["Stock Correlation & Portfolio Analytics Summary"])
        ws_summary.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        ws_summary.append([])

        metadata_rows = [
            ["Market Benchmark(s)", ", ".join(benchmarks)],
            ["Analysis Period", period],
            ["Filter Criteria", filter_label],
            ["Exported Assets Count", len(rows)],
            ["Export Timestamp (UTC)", datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")]
        ]

        # Calculate average correlations and betas per benchmark
        for bm in benchmarks:
            bm_corrs = []
            bm_betas = []
            for r in rows:
                c = clean_value(r.get(f"correlation_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("correlation") or r.get("correlation"))
                b = clean_value(r.get(f"beta_{bm}") or r.get("benchmarks_metrics", {}).get(bm, {}).get("beta") or r.get("beta"))
                if c is not None:
                    bm_corrs.append(c)
                if b is not None:
                    bm_betas.append(b)

            avg_c = float(np.mean(bm_corrs)) if bm_corrs else None
            avg_b = float(np.mean(bm_betas)) if bm_betas else None
            metadata_rows.append([f"Avg Correlation vs {bm}", f"{avg_c:.3f}" if avg_c is not None else "—"])
            metadata_rows.append([f"Avg Beta vs {bm}", f"{avg_b:.3f}" if avg_b is not None else "—"])

        meta_label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        meta_label_font = Font(name="Calibri", size=10, bold=True, color="334155")

        for row_data in metadata_rows:
            ws_summary.append(row_data)
            curr_row = ws_summary.max_row
            cell_lbl = ws_summary.cell(row=curr_row, column=1)
            cell_val = ws_summary.cell(row=curr_row, column=2)

            cell_lbl.fill = meta_label_fill
            cell_lbl.font = meta_label_font
            cell_lbl.border = cell_border

            cell_val.font = Font(name="Calibri", size=10)
            cell_val.border = cell_border

        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 32

        # Save workbook to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        clean_bm = "_".join("".join(c for c in bm if c.isalnum()) for bm in benchmarks[:3]) or "benchmark"
        clean_period = "".join(c for c in period if c.isalnum()) or "1y"
        filename = f"correlation_{clean_bm}_{clean_period}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("Excel export error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# DEDICATED CUSTOM MOVE & RELATIVE PERFORMANCE TRACKER
# ============================================================

@app.route("/api/calculate-custom-tracker", methods=["POST"])
def calculate_custom_tracker():
    try:
        data = request.get_json() or {}
        items = data.get("items", []) # list of dicts: [{"ticker": "COHR", "actual_move": 0.04}, ...]
        reference = str(data.get("reference") or data.get("benchmark") or "SPY").strip().upper()
        reference_move = data.get("reference_move", 0.01)
        period = str(data.get("period", "1y"))

        try:
            reference_move = float(reference_move)
        except (TypeError, ValueError):
            reference_move = 0.01

        if not items:
            return jsonify({"rows": [], "reference": reference, "reference_move": reference_move, "period": period})

        # Clean items & moves
        parsed_items = {}
        for item in items:
            if isinstance(item, dict):
                sym = str(item.get("ticker", "")).strip().upper()
                mv = item.get("actual_move")
            else:
                sym = str(item).strip().upper()
                mv = None
            if sym and sym != reference:
                try:
                    mv_val = float(mv) if mv is not None else None
                except (TypeError, ValueError):
                    mv_val = None
                parsed_items[sym] = mv_val

        tickers_list = list(parsed_items.keys())
        if not tickers_list:
            return jsonify({"rows": [], "reference": reference, "reference_move": reference_move, "period": period})

        all_symbols = [reference] + tickers_list
        df = get_stock_data(*all_symbols)

        if df is None or df.empty:
            return jsonify({"error": "Данные для указанных тикеров не найдены в базе данных."}), 404

        available_tickers = set(
            df["ticker"].astype(str).str.strip().str.upper().unique()
        )

        if reference not in available_tickers:
            return jsonify({"error": f"Базовый актив '{reference}' не найден в базе данных."}), 404

        valid_tickers = [t for t in tickers_list if t in available_tickers]
        if not valid_tickers:
            return jsonify({"error": "Ни один из указанных тикеров не найден в базе данных."}), 404

        analysis = calculate_groups_analysis(
            df=df,
            groups={"CustomTracker": valid_tickers},
            benchmarks=[reference],
            period=period,
            benchmark_move=reference_move,
            asset_moves={t: parsed_items[t] for t in valid_tickers if parsed_items[t] is not None}
        )

        ignored_tickers = analysis.get("ignored_tickers", {})
        partial_tickers = analysis.get("partial_tickers", {})
        bm_data = analysis.get("benchmarks_data", {}).get(reference, {})
        expected_range = bm_data.get("expected_range", {})
        correlation_matrix = analysis.get("correlation", {})
        analyzed_tickers = analysis.get("tickers", [])

        rows = []
        for ticker in analyzed_tickers:
            exp_info = expected_range.get(ticker, {})
            corr_val = correlation_matrix.get(ticker, {}).get(reference)
            act_mv = parsed_items.get(ticker)
            if act_mv is None and exp_info.get("actual_move") is not None:
                act_mv = exp_info.get("actual_move")

            rows.append({
                "ticker": ticker,
                "correlation": clean_value(corr_val),
                "beta": clean_value(exp_info.get("beta")),
                "expected_move": clean_value(exp_info.get("expected_stock_move")),
                "actual_move": clean_value(act_mv),
                "difference": clean_value(exp_info.get("difference")),
                "normalized": clean_value(exp_info.get("normalized_difference")),
                "volatility": clean_value(exp_info.get("volatility")),
                "residual_volatility": clean_value(exp_info.get("residual_volatility")),
                "observations": clean_value(exp_info.get("observations")),
                "reference": reference,
                "period": period,
                "is_partial_history": bool(exp_info.get("is_partial_history", False)),
                "history_period_ru": exp_info.get("history_period_ru"),
                "history_text": exp_info.get("history_text"),
                "history_start": exp_info.get("history_start"),
                "history_end": exp_info.get("history_end"),
                "history_days": clean_value(exp_info.get("history_days")),
                "history_observations": clean_value(exp_info.get("history_observations")),
                "selected_period": exp_info.get("selected_period", period),
                "selected_period_label": exp_info.get("selected_period_label"),
            })

        ignored_list = list(ignored_tickers.keys()) if isinstance(ignored_tickers, dict) else list(ignored_tickers)
        partial_list = list(partial_tickers.keys()) if isinstance(partial_tickers, dict) else list(partial_tickers)

        return jsonify({
            "rows": rows,
            "reference": reference,
            "reference_move": reference_move,
            "period": period,
            "missing": [t for t in tickers_list if t not in available_tickers],
            "ignored": ignored_list,
            "ignored_tickers": clean_dict(ignored_tickers),
            "partial": partial_list,
            "partial_tickers": clean_dict(partial_tickers),
        })

    except Exception as e:
        print("Custom tracker calculation error:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# DATABASE STATUS
# ============================================================

@app.route("/api/db-status")
def db_status():
    try:
        return jsonify(get_database_status())
    except Exception as e:
        return jsonify({
            "online": False,
            "source": "error",
            "message": str(e)
        }), 500


# ============================================================
# TICKERS LIST
# ============================================================

@app.route("/api/tickers")
def tickers():
    try:
        return jsonify(get_available_tickers())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# SERVER
# ============================================================

if __name__ == "__main__":
    app.run(
        host="127.0.0.1",
        port=8000,
        debug=True
    )